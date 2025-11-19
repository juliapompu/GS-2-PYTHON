import csv
import os
from typing import Dict, List, Optional

try:
    from openpyxl import load_workbook  # type: ignore
except ImportError:
    load_workbook = None


def ler_csv_lista_de_listas(caminho: str, delimitador: str = ";") -> List[List[str]]:
    tabela = []
    with open(caminho, "r", encoding="latin-1") as f:
        linhas = f.read().splitlines()
    if len(linhas) == 1 or all((delimitador in linha) for linha in linhas):
        tabela = [linha.split(delimitador) for linha in linhas]
    else:
        with open(caminho, newline="", encoding="latin-1") as f:
            leitor = csv.reader(f, delimiter=delimitador)
            tabela = list(leitor)

    return tabela


def ler_excel_lista_de_listas(caminho: str, aba: Optional[str] = None) -> List[List[str]]:
    if load_workbook is None:
        raise RuntimeError(
            "A biblioteca openpyxl não está instalada. "
            "Instale-a com 'pip install openpyxl' para ler arquivos Excel."
        )

    workbook = load_workbook(filename=caminho, data_only=True, read_only=True)
    if aba:
        if aba not in workbook.sheetnames:
            raise ValueError(f"Aba '{aba}' não encontrada. Abas disponíveis: {workbook.sheetnames}")
        worksheet = workbook[aba]
    else:
        worksheet = workbook.active

    tabela = []
    for linha in worksheet.iter_rows(values_only=True):
        if linha is None:
            continue
        tabela.append([(str(celda).strip() if celda is not None else "").strip() for celda in linha])

    return tabela


def carregar_tabela(caminho: str, aba: Optional[str] = None, delimitador: str = ";") -> List[List[str]]:
    if not os.path.exists(caminho):
        raise FileNotFoundError(f"Arquivo '{caminho}' não encontrado.")

    _, ext = os.path.splitext(caminho.lower())
    if ext in {".csv", ".txt"}:
        return ler_csv_lista_de_listas(caminho, delimitador)
    if ext in {".xlsx", ".xlsm", ".xls"}:
        return ler_excel_lista_de_listas(caminho, aba)

    raise ValueError("Formato de arquivo não suportado. Utilize CSV ou Excel.")


def lista_para_dict(lista):
    cabecalho = [c.strip() for c in lista[0]]
    dados = []

    for linha in lista[1:]:
        d = {}
        for i, campo in enumerate(cabecalho):
            valor = linha[i].strip() if i < len(linha) else ""
            d[campo] = valor
        dados.append(d)

    return dados


def organizar_por_pais(lista_dic):
    colunas_pais_possiveis = ["Pais", "PAIS", "pais", "País", "país"]
    chave_pais = None

    for chave in colunas_pais_possiveis:
        if chave in lista_dic[0]:
            chave_pais = chave
            break

    if chave_pais is None:
        raise KeyError(
            "Nenhuma coluna de país encontrada no arquivo! "
            f"Colunas disponíveis: {list(lista_dic[0].keys())}"
        )

    dados = {}
    for linha in lista_dic:
        pais = linha[chave_pais]
        dados[pais] = linha
    return dados


def normalizar_numero(texto: str) -> Optional[float]:
    if texto is None:
        return None
    bruto = texto.strip().replace(" ", "")
    if bruto == "" or bruto in {"-", "--", "None"}:
        return None

    if "," in bruto and "." in bruto:
        bruto = bruto.replace(".", "").replace(",", ".")
    elif "," in bruto:
        if bruto.count(",") == 1:
            bruto = bruto.replace(",", ".")
        else:
            bruto = bruto.replace(",", "")
    else:
        bruto = bruto.replace(",", "")

    try:
        return float(bruto)
    except ValueError:
        return None


def apresenta_dado(dados, nome_dado):
    resultado = {}
    for pais, valores in dados.items():
        if nome_dado in valores:
            valor = valores[nome_dado].strip()
            if valor in {"", "-", "--"}:
                resultado[pais] = None
            else:
                resultado[pais] = valor
    return resultado


def valores_numericos(dados, nome_dado) -> List[float]:
    valores = []
    for valor in apresenta_dado(dados, nome_dado).values():
        numero = normalizar_numero(valor) if isinstance(valor, str) else valor
        if numero is not None:
            valores.append(numero)
    return valores


def apresenta_pais(dados, nome_pais):
    return dados.get(nome_pais, None)


def media_dado(dados, nome_dado):
    valores = valores_numericos(dados, nome_dado)
    if not valores:
        return None
    return sum(valores) / len(valores)


def variancia_dado(dados, nome_dado):
    valores = valores_numericos(dados, nome_dado)
    if not valores:
        return None
    m = sum(valores) / len(valores)
    return sum((v - m) ** 2 for v in valores) / len(valores)


def media_ponderada(dados, nome_dado, nome_peso):
    pares = []
    for pais, info in dados.items():
        try:
            x = normalizar_numero(info[nome_dado])
            w = normalizar_numero(info[nome_peso])
            if x is not None and w is not None:
                pares.append((x, w))
        except KeyError:
            pass

    numerador = sum(x * w for x, w in pares)
    denominador = sum(w for _, w in pares)
    return numerador / denominador if denominador != 0 else None


def mediana_dado(dados, nome_dado):
    valores = sorted(valores_numericos(dados, nome_dado))
    n = len(valores)

    if n == 0:
        return None
    if n % 2 == 1:
        return valores[n // 2]
    return (valores[n // 2 - 1] + valores[n // 2]) / 2


def colunas_disponiveis(dados: Dict[str, Dict[str, str]]) -> List[str]:
    if not dados:
        return []
    primeira_linha = next(iter(dados.values()))
    return list(primeira_linha.keys())


def solicitar_coluna(dados, mensagem: str) -> Optional[str]:
    coluna = input(mensagem).strip()
    if not coluna:
        print("Informe um nome de coluna válido.")
        return None
    if coluna not in colunas_disponiveis(dados):
        print(f"Coluna '{coluna}' não encontrada. Utilize a opção 7 para listar as colunas.")
        return None
    return coluna


def menu(dados):
    while True:
        print("\n===== MENU SAFE-DADOS =====")
        print("1 - Apresentar dado (ex: PIB, Populacao)")
        print("2 - Apresentar país (todos os dados)")
        print("3 - Média de um dado")
        print("4 - Variância de um dado")
        print("5 - Média ponderada")
        print("6 - Mediana de um dado")
        print("7 - Listar colunas disponíveis")
        print("0 - Sair")

        op = input("\nEscolha uma opção: ").strip()

        if op == "1":
            coluna = solicitar_coluna(dados, "Nome do dado: ")
            if coluna:
                print(apresenta_dado(dados, coluna))

        elif op == "2":
            pais = input("Nome do país: ").strip()
            info = apresenta_pais(dados, pais)
            if info:
                print(info)
            else:
                print("País não encontrado!")

        elif op == "3":
            coluna = solicitar_coluna(dados, "Dado para média: ")
            if coluna:
                resultado = media_dado(dados, coluna)
                print("Média =", resultado if resultado is not None else "Sem valores numéricos.")

        elif op == "4":
            print("Variância indica quanto os valores se afastam da média. "
                  "Use uma coluna numérica para ver a dispersão dos dados.")
            coluna = solicitar_coluna(dados, "Dado para variância: ")
            if coluna:
                resultado = variancia_dado(dados, coluna)
                print("Variância =", resultado if resultado is not None else "Sem valores numéricos.")

        elif op == "5":
            print("A média ponderada usa pesos (por exemplo população) para dar "
                  "mais importância a certos valores. Informe a coluna principal "
                  "e depois a coluna de pesos.")
            dado = solicitar_coluna(dados, "Dado principal: ")
            if not dado:
                continue
            peso = solicitar_coluna(dados, "Coluna de peso (ex: Populacao): ")
            if not peso:
                continue
            resultado = media_ponderada(dados, dado, peso)
            print("Média ponderada =", resultado if resultado is not None else "Sem valores válidos.")

        elif op == "6":
            print("A mediana é o valor central após ordenar os dados, útil para "
                  "evitar influência de valores extremos.")
            coluna = solicitar_coluna(dados, "Dado para mediana: ")
            if coluna:
                resultado = mediana_dado(dados, coluna)
                print("Mediana =", resultado if resultado is not None else "Sem valores numéricos.")

        elif op == "7":
            cols = colunas_disponiveis(dados)
            print("Colunas:", ", ".join(cols) if cols else "Nenhuma coluna encontrada.")

        elif op == "0":
            print("Saindo...")
            break

        else:
            print("Opção inválida!")


def executar_interface():
    caminho = input("Informe o caminho do arquivo CSV ou Excel (.csv/.xlsx): ").strip()
    if not caminho:
        caminho = os.path.join(os.path.dirname(__file__), "excel_python.csv")
        print(f"Nenhum caminho informado. Utilizando arquivo padrão: {caminho}")

    aba = None
    if caminho.lower().endswith((".xlsx", ".xlsm", ".xls")):
        aba = input("Nome da aba (pressione Enter para usar a primeira): ").strip() or None

    tabela = carregar_tabela(caminho, aba=aba)
    lista_dic = lista_para_dict(tabela)
    dados = organizar_por_pais(lista_dic)
    menu(dados)


if __name__ == "__main__":
    executar_interface()
